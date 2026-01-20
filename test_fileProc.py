import os
import tempfile
import shutil
import pytest
from openpyxl import load_workbook
from fileProc import filenames_to_excel


class TestFilenamesToExcel:
    """Unit tests for the filenames_to_excel function."""

    def setup_method(self):
        """Set up temporary directories for each test."""
        self.temp_dir = tempfile.mkdtemp()
        self.source_folder = os.path.join(self.temp_dir, "source")
        self.archive_folder = os.path.join(self.temp_dir, "archive")
        self.output_file = os.path.join(self.temp_dir, "output.xlsx")
        os.makedirs(self.source_folder)
        os.makedirs(self.archive_folder)

    def teardown_method(self):
        """Clean up temporary directories after each test."""
        shutil.rmtree(self.temp_dir)

    def create_test_file(self, filename):
        """Helper to create a test file in the source folder."""
        file_path = os.path.join(self.source_folder, filename)
        with open(file_path, "w") as f:
            f.write("test content")
        return file_path

    def test_file_with_exactly_four_parts(self):
        """Test processing a file with exactly 4 underscore-separated parts."""
        self.create_test_file("part1_part2_part3_part4.txt")

        filenames_to_excel(self.source_folder, self.output_file, self.archive_folder)

        wb = load_workbook(self.output_file)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        assert rows[0] == ("Col1", "Col2", "Col3", "Col4")
        assert rows[1] == ("part1", "part2", "part3", "part4")

    def test_file_with_fewer_than_four_parts(self):
        """Test processing a file with fewer than 4 parts (should pad with empty strings)."""
        self.create_test_file("part1_part2.txt")

        filenames_to_excel(self.source_folder, self.output_file, self.archive_folder)

        wb = load_workbook(self.output_file)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        assert rows[0] == ("Col1", "Col2", "Col3", "Col4")
        assert rows[1] == ("part1", "part2", None, None)

    def test_file_with_more_than_four_parts(self):
        """Test processing a file with more than 4 parts (should truncate to 4)."""
        self.create_test_file("part1_part2_part3_part4_part5_part6.txt")

        filenames_to_excel(self.source_folder, self.output_file, self.archive_folder)

        wb = load_workbook(self.output_file)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        assert rows[0] == ("Col1", "Col2", "Col3", "Col4")
        assert rows[1] == ("part1", "part2", "part3", "part4")

    def test_file_with_single_part(self):
        """Test processing a file with no underscores (single part)."""
        self.create_test_file("singlepart.txt")

        filenames_to_excel(self.source_folder, self.output_file, self.archive_folder)

        wb = load_workbook(self.output_file)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        assert rows[0] == ("Col1", "Col2", "Col3", "Col4")
        assert rows[1] == ("singlepart", None, None, None)

    def test_empty_folder(self):
        """Test processing an empty folder (should create Excel with only headers)."""
        filenames_to_excel(self.source_folder, self.output_file, self.archive_folder)

        wb = load_workbook(self.output_file)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        assert len(rows) == 1
        assert rows[0] == ("Col1", "Col2", "Col3", "Col4")

    def test_files_moved_to_archive(self):
        """Test that processed files are moved to the archive folder."""
        self.create_test_file("test_file.txt")

        filenames_to_excel(self.source_folder, self.output_file, self.archive_folder)

        assert not os.path.exists(os.path.join(self.source_folder, "test_file.txt"))
        assert os.path.exists(os.path.join(self.archive_folder, "test_file.txt"))

    def test_excel_file_created(self):
        """Test that the Excel file is created at the specified path."""
        self.create_test_file("test_file.txt")

        filenames_to_excel(self.source_folder, self.output_file, self.archive_folder)

        assert os.path.exists(self.output_file)

    def test_worksheet_title(self):
        """Test that the worksheet has the correct title."""
        self.create_test_file("test_file.txt")

        filenames_to_excel(self.source_folder, self.output_file, self.archive_folder)

        wb = load_workbook(self.output_file)
        ws = wb.active
        assert ws.title == "FileWords"

    def test_multiple_files(self):
        """Test processing multiple files at once."""
        self.create_test_file("a_b_c_d.txt")
        self.create_test_file("e_f.txt")
        self.create_test_file("single.txt")

        filenames_to_excel(self.source_folder, self.output_file, self.archive_folder)

        wb = load_workbook(self.output_file)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        assert rows[0] == ("Col1", "Col2", "Col3", "Col4")
        assert len(rows) == 4

        data_rows = sorted(rows[1:], key=lambda x: x[0] if x[0] else "")
        assert ("a", "b", "c", "d") in data_rows
        assert ("e", "f", None, None) in data_rows
        assert ("single", None, None, None) in data_rows

    def test_ignores_subdirectories(self):
        """Test that subdirectories in the source folder are ignored."""
        self.create_test_file("test_file.txt")
        subdir = os.path.join(self.source_folder, "subdir_with_underscores")
        os.makedirs(subdir)

        filenames_to_excel(self.source_folder, self.output_file, self.archive_folder)

        wb = load_workbook(self.output_file)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        assert len(rows) == 2
        assert rows[1] == ("test", "file", None, None)

    def test_file_with_multiple_extensions(self):
        """Test processing a file with multiple dots in the name."""
        self.create_test_file("part1_part2.backup.txt")

        filenames_to_excel(self.source_folder, self.output_file, self.archive_folder)

        wb = load_workbook(self.output_file)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        assert rows[1] == ("part1", "part2.backup", None, None)
